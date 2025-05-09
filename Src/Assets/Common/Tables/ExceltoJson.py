import json
import openpyxl
from pathlib import Path
import sys
import ast
import os
from typing import Union

def parse_string_to_array(value: str) -> Union[list, str]:
    """安全地将字符串数组转换为Python列表"""
    if not isinstance(value, str):
        return value
    
    value = value.strip()
    
    # 检查是否是数组格式
    if value.startswith('[') and value.endswith(']'):
        try:
            # 使用json.loads解析（要求严格JSON格式）
            return json.loads(value)
        except json.JSONDecodeError:
            try:
                # 尝试用ast.literal_eval作为备用方案
                parsed = ast.literal_eval(value)
                return parsed if isinstance(parsed, list) else value
            except (ValueError, SyntaxError):
                return value
    return value

def excel_to_json(input_excel_path: Union[str, Path], 
                 output_dir_path: Union[str, Path], 
                 start_row: int = 2, 
                 end_row: int = 0) -> bool:
    """
    转换Excel所有Sheet到JSON文件（第一行为列名）
    
    参数:
        input_excel_path: 输入的Excel文件路径
        output_dir_path: JSON输出目录
        start_row: 数据起始行（默认第2行，第1行为列名）
        end_row: 数据结束行（0表示所有行）
    
    返回:
        bool: 是否转换成功
    """
    try:
        # 路径标准化处理
        input_excel_path = Path(input_excel_path).absolute()
        output_dir = Path(output_dir_path).absolute()
        
        # 创建以Excel文件名命名的子目录
        excel_name = input_excel_path.stem
        sheet_output_dir = output_dir / excel_name
        sheet_output_dir.mkdir(parents=True, exist_ok=True)
        
        print(f"\n处理文件: {input_excel_path.name}")
        print(f"输出目录: {sheet_output_dir}")

        # 加载工作簿
        workbook = openpyxl.load_workbook(input_excel_path, data_only=True)
        
        for sheet in workbook.worksheets:
            # 跳过空工作表
            if sheet.max_row == 0:
                print(f"  跳过空工作表: {sheet.title}")
                continue
                
            data = []
            
            # 获取列名（第一行）
            headers = []
            for cell in sheet[1]:
                header = str(cell.value) if cell.value is not None else f"Column_{cell.column}"
                headers.append(header)
            
            # 确定数据行范围
            data_start_row = max(2, start_row)  # 数据从第2行开始
            data_end_row = end_row if end_row > 0 else sheet.max_row
            
            # 读取数据行
            for row in sheet.iter_rows(
                min_row=data_start_row,
                max_row=data_end_row,
                values_only=True
            ):
                if any(cell is not None for cell in row):
                    row_dict = {}
                    for header, cell in zip(headers, row):
                        if cell is None:
                            continue
                        # 处理数组数据
                        processed_value = parse_string_to_array(str(cell)) if isinstance(cell, (str, bytes)) else cell
                        row_dict[header] = processed_value
                    
                    data.append(row_dict)
            
            # 构建JSON文件路径
            json_filename = f"{clean_filename(sheet.title)}.json"
            json_path = sheet_output_dir / json_filename
            
            # 写入JSON文件
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=4)
            
            print(f"  ✓ 工作表 [{sheet.title}] -> {json_filename}")
        
        return True
    
    except Exception as e:
        print(f"✗ 转换失败 [{input_excel_path.name}]: {str(e)}")
        return False

def excel_to_json_2dArray(input_excel_path: Union[str, Path], 
                 output_dir_path: Union[str, Path], ) -> bool:
    """
    转换Excel所有Sheet到JSON文件：以简单的二维数组形式转化
    
    参数:
        input_excel_path: 输入的Excel文件路径
        output_dir_path: JSON输出目录
        路径为相对路径
    
    返回:
        bool: 是否转换成功
    """
    try:
        # 路径标准化处理
        input_excel_path = Path(input_excel_path).absolute()
        output_dir = Path(output_dir_path).absolute()
        
        # 创建以Excel文件名命名的子目录
        excel_name = input_excel_path.stem
        sheet_output_dir = output_dir / excel_name
        sheet_output_dir.mkdir(parents=True, exist_ok=True)
        
        print(f"\n处理文件: {input_excel_path.name}")
        print(f"输出目录: {sheet_output_dir}")

        # 加载工作簿
        workbook = openpyxl.load_workbook(input_excel_path, data_only=True)
        
        for sheet in workbook.worksheets:
            # 跳过空工作表
            if sheet.max_row == 0:
                print(f"  跳过空工作表: {sheet.title}")
                continue
                
            data = [[cell if cell is not None else 0 for cell in row] for row in sheet.values]
            

             # 构建JSON文件路径
            json_filename = f"{clean_filename(sheet.title)}.json"
            json_path = sheet_output_dir / json_filename
            
            # 写入JSON文件
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=4)
            
            print(f"  ✓ 工作表 [{sheet.title}] -> {json_filename}")
        return True
    except Exception as e:
        print(f"✗ 转换失败 [{input_excel_path.name}]: {str(e)}")
        return False

def excel_to_json_Dictionary(input_excel_path: Union[str, Path], 
                 output_dir_path: Union[str, Path], 
                 start_row: int = 2, 
                 end_row: int = 0) -> bool:
    """
    转换Excel所有Sheet到JSON文件（第一行为列名）
    
    参数:
        input_excel_path: 输入的Excel文件路径
        output_dir_path: JSON输出目录
        start_row: 数据起始行（默认第2行，第1行为列名）
        end_row: 数据结束行（0表示所有行）
    
    返回:
        bool: 是否转换成功
    """
    try:
        # 路径标准化处理
        input_excel_path = Path(input_excel_path).absolute()
        output_dir = Path(output_dir_path).absolute()
        
        # 创建以Excel文件名命名的子目录
        excel_name = input_excel_path.stem
        sheet_output_dir = output_dir / excel_name
        sheet_output_dir.mkdir(parents=True, exist_ok=True)
        
        print(f"\n处理文件: {input_excel_path.name}")
        print(f"输出目录: {sheet_output_dir}")

        # 加载工作簿
        workbook = openpyxl.load_workbook(input_excel_path, data_only=True)
        
        for sheet in workbook.worksheets:
            # 跳过空工作表
            if sheet.max_row == 0:
                print(f"  跳过空工作表: {sheet.title}")
                continue
                
            #字典结构的result_dict
            result_dict={}
            
            # 获取列名（第一行）
            headers = []
            for cell in sheet[1]:
                header = str(cell.value) if cell.value is not None else f"Column_{cell.column}"
                headers.append(header)
            try:
                key_col_index = headers.index("Key")
            except ValueError:
                print(f"  工作表 {sheet.title} 缺少Key列，已跳过")
                return
            
            # 确定数据行范围
            data_start_row = max(2, start_row)  # 数据从第2行开始
            data_end_row = end_row if end_row > 0 else sheet.max_row
            
            # 读取数据行
            for row in sheet.iter_rows(
                min_row=data_start_row,
                max_row=data_end_row,
                values_only=True
            ):
                # 获取 key 并保留原始类型（不强制转 int）
                raw_key = row[key_col_index]
                if raw_key is None:
                    continue  # 跳过空 key
                
                # 确定 key 的类型（如果是数字字符串，则转为 int；否则保留 string）
                try:
                    key = int(raw_key) if isinstance(raw_key, str) and raw_key.isdigit() else raw_key
                except (ValueError, TypeError):
                    key = raw_key  # 如果转换失败，保持原
                if any(cell is not None for cell in row):
                    row_dict = {}
                    for header, cell in zip(headers, row):
                        if cell is None:
                            continue
                        if header == "Key":
                            continue
                        # 处理数组数据
                        processed_value = parse_string_to_array(str(cell)) if isinstance(cell, (str, bytes)) else cell
                        row_dict[header] = processed_value
                result_dict[key] = row_dict
            
            # 构建JSON文件路径
            json_filename = f"{clean_filename(sheet.title)}.json"
            json_path = sheet_output_dir / json_filename
            
            # 写入JSON文件
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(result_dict, f, ensure_ascii=False, indent=4)
            
            print(f"  ✓ 工作表 [{sheet.title}] -> {json_filename}")
        
        return True
    
    except Exception as e:
        print(f"✗ 转换失败 [{input_excel_path.name}]: {str(e)}")
        return False

def clean_filename(filename: str) -> str:
    """清理文件名中的非法字符"""
    invalid_chars = '<>:"/\\|?*'
    for char in invalid_chars:
        filename = filename.replace(char, '_')
    return filename

def batch_convert(input_dir: Union[str, Path], 
                 output_dir: Union[str, Path], 
                 start_row: int, 
                 end_row: int) -> bool:
    """
    批量转换目录下的所有Excel文件
    
    参数:
        input_dir: 输入目录路径
        output_dir: 输出目录路径
        start_row: 起始行号
        end_row: 结束行号
    """
    input_dir = Path(input_dir).absolute()
    output_dir = Path(output_dir).absolute()
    
    if not input_dir.exists():
        print(f"错误: 输入目录不存在 {input_dir}")
        return False
    
    # 创建输出目录
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # 查找所有Excel文件
    excel_files = list(input_dir.glob('*.xlsx')) + list(input_dir.glob('*.xls'))
    if not excel_files:
        print(f"警告: 没有找到Excel文件 {input_dir}")
        return False
    
    print(f"发现 {len(excel_files)} 个Excel文件")
    success_count = 0
    
    for excel_file in excel_files:
        if excel_to_json(excel_file, output_dir, start_row, end_row):
            success_count += 1
    
    print(f"\n转换完成: 成功 {success_count}/{len(excel_files)} 个文件")
    return success_count > 0

def generate_cmd_script(script_dir: Union[str, Path],
                       input_dir: Union[str, Path],
                       output_dir: Union[str, Path],
                       start_row: int,
                       end_row: int) -> Path:
    """
    生成可重复运行的CMD脚本：我懒得维护了
    
    返回:
        Path: 生成的CMD文件路径
    """
    def get_relative(target: Path) -> str:
        """获取相对于脚本目录的路径"""
        try:
            rel_path = target.relative_to(script_dir)
            return f".\\{rel_path}" if not str(rel_path).startswith("..") else f"{rel_path}"
        except ValueError:
            return str(target)
    
    script_dir = Path(script_dir).absolute()
    rel_input = get_relative(Path(input_dir).absolute())
    rel_output = get_relative(Path(output_dir).absolute())
    rel_script = get_relative(Path(__file__).absolute())

    cmd_content = f"""@echo off
chcp 65001 > nul
echo 批量Excel转JSON工具（第一行为列名）
echo 输入目录: {rel_input}
echo 输出目录: {rel_output}

python "{rel_script}" ^
"{rel_input}" ^
"{rel_output}" ^
{start_row} ^
{end_row}

if %errorlevel% equ 0 (
    echo 所有文件转换完成!
) else (
    echo 部分文件转换失败
)
pause
"""
    cmd_path = script_dir / "excel_to_json_batch.cmd"
    cmd_path.write_text(cmd_content, encoding='utf-8')
    print(f"\n✓ 生成的CMD文件: {cmd_path}")
    return cmd_path

def main():
    # 确保在脚本目录执行
    os.chdir(Path(__file__).parent)
    
    Useby = input("交互模式？Y：1，N：0").strip('"')
    if Useby == 1:
        print("=== Excel转JSON工具（第一行为列名）===")
        
        current_dir = Path(__file__).parent.absolute()
        
        # 输入输出目录设置
        input_dir = input("Excel输入目录路径（回车使用Tables）: ").strip('"') or "Tables"
        output_dir = input("JSON输出目录路径（回车使用Data）: ").strip('"') or "Data"
        script_dir = input("脚本工作目录路径（回车使用当前目录）: ").strip('"') or "."
        
        # 行范围设置
        while True:
            try:
                start_row = int(input("数据起始行号（从2开始，回车默认2）: ") or "2")
                end_row = int(input("数据结束行号（0表示所有行，回车默认0）: ") or "0")
                
                if start_row < 2:
                    print("错误: 数据起始行必须≥2（第1行是列名）!")
                    continue
                if end_row > 0 and end_row < start_row:
                    print("错误: 结束行号不能小于起始行号!")
                    continue
                break
            except ValueError:
                print("错误: 请输入有效数字!")
        if(os.path.isdir(input_dir)):
            excel_to_json(input_dir,output_dir,start_row,end_row)
        else:
            # 创建目录
            Path(input_dir).mkdir(parents=True, exist_ok=True)
            Path(output_dir).mkdir(parents=True, exist_ok=True)
            Path(script_dir).mkdir(parents=True, exist_ok=True)

            # 执行转换
            batch_convert(input_dir,output_dir,start_row, end_row)
    else:
        # 默认转换逻辑
        excel_to_json_Dictionary("Tables/Space.xlsx","Data")
        excel_to_json_Dictionary("Tables/ActionDefine.xlsx","Data")
        excel_to_json_2dArray("Tables/Versus.xlsx","Data")


if __name__ == "__main__":
    main()
